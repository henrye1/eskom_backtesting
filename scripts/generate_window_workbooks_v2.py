"""Generate per-window workbooks with FULL calculation sheets matching the
reference workbook structure (Munic_dashboard_LPU_1.xlsx).

Each vintage calculation sheet includes:
  - Balance triangle (masked to observable data at that vintage date)
  - Aggregate recoveries vector
  - Cumulative balance matrix
  - Discount factor matrix
  - LGD component matrix (Recovery/CumBal * DF)
  - LGD term structure (= 1 - column sums of component matrix)
  - Weighted LGD
  - Parameters block (interest rate, window, min TID)

This provides a complete audit trail linking the methodology document
to the actual calculations in each workbook.
"""

import os
import sys
import warnings

import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from lgd_model.config import ModelConfig
from lgd_model.data_loader import load_recovery_triangle, extract_balance_matrix
from lgd_model.core_engine import (
    compute_aggregate_recoveries,
    compute_cumulative_balances,
    compute_discount_matrix,
    compute_lgd_term_structure,
    compute_ead_weighted_lgd,
)
from lgd_model.vintage import run_vintage_analysis
from lgd_model.backtest import run_backtest
from lgd_model.scenario import run_scenario

# ── Formatting constants ──────────────────────────────────────────────
HDR_FILL = PatternFill('solid', fgColor='2E75B6')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=9)
SECTION_FONT = Font(name='Arial', bold=True, size=10, color='2E75B6')
LABEL_FONT = Font(name='Arial', bold=True, size=9)
DATA_FONT = Font(name='Arial', size=9)
PARAM_FONT = Font(name='Arial', bold=True, size=9, color='333333')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
PCT_FMT = '0.0000%'
NUM_FMT = '#,##0.00'
DATE_FMT = 'YYYY-MM-DD'


def _style_hdr(cell):
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal='center')
    cell.border = THIN_BORDER


def _style_data(cell, fmt=None):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')
    if fmt:
        cell.number_format = fmt


def _write_val(ws, row, col, val, fmt=None, font=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = font or DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')
    if fmt:
        cell.number_format = fmt
    return cell


def _write_matrix(ws, start_row, start_col, matrix, n_rows, n_cols, fmt=None):
    """Write a 2D numpy array to the sheet."""
    for r in range(n_rows):
        for c in range(n_cols):
            val = matrix[r, c] if r < matrix.shape[0] and c < matrix.shape[1] else np.nan
            cell = ws.cell(row=start_row + r, column=start_col + c)
            if np.isnan(val):
                cell.value = None
            else:
                cell.value = float(val)
            _style_data(cell, fmt)


def write_vintage_calc_sheet(wb, vintage_label, balance_window, eads_window,
                             master_tids_window, n_periods, config,
                             recoveries, cum_bal, discount_mat, lgd_ts,
                             weighted_lgd, cohort_indices):
    """Write a single vintage calculation sheet with full audit trail.

    Layout matches the reference workbook:
    - Cols A-D: Cohort index, max TID, LGD, EAD
    - Col F: TID index
    - Col G: MIN balance
    - Col H onwards: Balance triangle
    - Row (n_cohorts+4): Weighted LGD + Check row
    - Row (n_cohorts+6): Parameters
    - Row (n_cohorts+11): Recovery vector
    - Row (n_cohorts+13)+: Cumulative balance matrix
    - Row (n_cohorts+13+n_periods+3)+: Discount matrix
    - Row (n_cohorts+13+2*(n_periods+3))+: LGD component matrix
    """
    sheet_label = vintage_label.replace("Latest ", "")
    sheet_name = f'RR LGD TERM STRUCTURES  ({sheet_label})'
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    ws = wb.create_sheet(sheet_name)
    n_cohorts = balance_window.shape[0]

    # ── Row 1: Header ──
    _write_val(ws, 1, 3, 'LGD TERM STRUCTURES', font=SECTION_FONT)
    _write_val(ws, 1, 7, 'MIN', font=LABEL_FONT)
    for t in range(n_periods + 1):
        _write_val(ws, 1, 8 + t, t, font=HDR_FONT)
        ws.cell(row=1, column=8 + t).fill = HDR_FILL

    # ── Rows 2..(n_cohorts+1): Balance triangle with cohort metadata ──
    for i in range(n_cohorts):
        r = 2 + i
        _write_val(ws, r, 1, int(cohort_indices[i]))  # Cohort index
        max_tid = int(master_tids_window[i])
        _write_val(ws, r, 2, max_tid)  # Max TID for this cohort

        # LGD for this cohort at TID 0
        lgd_val = float(lgd_ts[0]) if len(lgd_ts) > 0 else None
        _write_val(ws, r, 3, float(lgd_ts[min(i, len(lgd_ts) - 1)]) if i < len(lgd_ts) else 1.0, fmt=PCT_FMT)

        # EAD
        _write_val(ws, r, 4, float(eads_window[i]), fmt=NUM_FMT)

        # TID index
        _write_val(ws, r, 6, i)

        # MIN balance (minimum non-NaN value in this cohort's row)
        row_vals = balance_window[i, :]
        valid_vals = row_vals[~np.isnan(row_vals)]
        if len(valid_vals) > 0:
            _write_val(ws, r, 7, float(np.min(valid_vals)), fmt=NUM_FMT)

        # Balance triangle values
        for t in range(n_periods):
            val = balance_window[i, t]
            cell = ws.cell(row=r, column=8 + t)
            if np.isnan(val):
                cell.value = None
            else:
                cell.value = float(val)
            _style_data(cell, NUM_FMT)

    # ── Weighted LGD + Check row ──
    check_row = 2 + n_cohorts + 2
    _write_val(ws, check_row, 1, 'Weighted', font=LABEL_FONT)
    _write_val(ws, check_row, 3, float(weighted_lgd), fmt=PCT_FMT)
    _write_val(ws, check_row, 7, 'Check', font=LABEL_FONT)
    for t in range(min(n_periods + 1, len(lgd_ts))):
        _write_val(ws, check_row, 8 + t, float(lgd_ts[t]), fmt=PCT_FMT)

    # ── Parameters block ──
    param_row = check_row + 2
    _write_val(ws, param_row, 7, 'Interest Rate', font=PARAM_FONT)
    _write_val(ws, param_row, 8, config.discount_rate)
    _write_val(ws, param_row + 1, 7, 'Columns', font=PARAM_FONT)
    _write_val(ws, param_row + 1, 8, n_periods)
    _write_val(ws, param_row + 2, 7, 'Rows', font=PARAM_FONT)
    _write_val(ws, param_row + 2, 8, n_cohorts)
    _write_val(ws, param_row + 3, 7, 'Min', font=PARAM_FONT)
    _write_val(ws, param_row + 3, 8, 0)

    # ── Recovery vector ──
    recov_row = param_row + 5
    _write_val(ws, recov_row, 7, 'Recovery', font=SECTION_FONT)
    for t in range(n_periods):
        _write_val(ws, recov_row, 8 + t, float(recoveries[t]), fmt=NUM_FMT)

    # ── Cumulative Balance matrix ──
    cumbal_row = recov_row + 2
    for t_row in range(n_periods + 1):
        _write_val(ws, cumbal_row + t_row, 7, t_row)
        for t_col in range(n_periods):
            if t_row < cum_bal.shape[0] and t_col < cum_bal.shape[1]:
                val = cum_bal[t_row, t_col]
                cell = ws.cell(row=cumbal_row + t_row, column=8 + t_col)
                if np.isnan(val) or (t_col < t_row):
                    cell.value = None
                else:
                    cell.value = float(val)
                _style_data(cell, NUM_FMT)

    # ── Discount matrix ──
    disc_row = cumbal_row + n_periods + 3
    # TID header row
    for t in range(n_periods + 1):
        _write_val(ws, disc_row - 1, 8 + t, t)

    for t_row in range(n_periods + 1):
        _write_val(ws, disc_row + t_row, 7, t_row)
        for t_col in range(n_periods):
            if t_row < discount_mat.shape[0] and t_col < discount_mat.shape[1]:
                val = discount_mat[t_row, t_col]
                cell = ws.cell(row=disc_row + t_row, column=8 + t_col)
                if t_col < t_row:
                    cell.value = None
                else:
                    cell.value = float(val)
                _style_data(cell, '0.000000')

    # ── LGD component matrix (Recovery[c] / CumBal[t,c] * DF[t,c]) ──
    lgd_comp_row = disc_row + n_periods + 3
    for t_row in range(n_periods + 1):
        _write_val(ws, lgd_comp_row + t_row, 7, t_row)
        for t_col in range(n_periods):
            cell = ws.cell(row=lgd_comp_row + t_row, column=8 + t_col)
            if t_col < t_row:
                cell.value = None
            elif (t_row < cum_bal.shape[0] and t_col < cum_bal.shape[1]
                  and not np.isnan(cum_bal[t_row, t_col])
                  and cum_bal[t_row, t_col] != 0
                  and t_col < len(recoveries)
                  and t_row < discount_mat.shape[0]
                  and t_col < discount_mat.shape[1]):
                comp = (recoveries[t_col] / cum_bal[t_row, t_col]) * discount_mat[t_row, t_col]
                cell.value = float(comp)
            else:
                cell.value = None
            _style_data(cell, PCT_FMT)

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['F'].width = 6
    ws.column_dimensions['G'].width = 14

    return ws


def write_backtest_section(ws, start_row, title, vintage_labels, periods,
                           hindsights, matrix, n_tid):
    """Write a labelled backtest matrix section."""
    ws.cell(row=start_row, column=1, value=title).font = SECTION_FONT
    hdr_row = start_row + 1
    _write_val(ws, hdr_row, 1, 'Vintage', font=HDR_FONT)
    ws.cell(row=hdr_row, column=1).fill = HDR_FILL
    _write_val(ws, hdr_row, 2, 'Period', font=HDR_FONT)
    ws.cell(row=hdr_row, column=2).fill = HDR_FILL
    _write_val(ws, hdr_row, 3, 'Hindsight', font=HDR_FONT)
    ws.cell(row=hdr_row, column=3).fill = HDR_FILL
    for t in range(n_tid):
        cell = ws.cell(row=hdr_row, column=4 + t, value=t)
        _style_hdr(cell)

    n_v = len(vintage_labels)
    for i in range(n_v):
        r = hdr_row + 1 + i
        _write_val(ws, r, 1, vintage_labels[i])
        period_val = periods[i]
        if hasattr(period_val, 'to_pydatetime'):
            period_val = period_val.to_pydatetime()
        _write_val(ws, r, 2, period_val, fmt=DATE_FMT)
        _write_val(ws, r, 3, hindsights[i])
        for t in range(min(n_tid, matrix.shape[1])):
            val = matrix[i, t]
            cell = ws.cell(row=r, column=4 + t)
            cell.value = None if np.isnan(val) else float(val)
            _style_data(cell, PCT_FMT)

    return hdr_row + 1 + n_v + 1


def copy_master_sheet(source_path, dest_wb):
    """Copy the RR LGD TERM STRUCTURE ALL sheet."""
    from copy import copy
    src_wb = load_workbook(source_path, data_only=True)
    src_ws = src_wb['RR LGD TERM STRUCTURE ALL']
    dest_ws = dest_wb.create_sheet('RR LGD TERM STRUCTURE ALL')
    for row in src_ws.iter_rows():
        for cell in row:
            dest_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dest_cell.font = copy(cell.font)
                dest_cell.fill = copy(cell.fill)
                dest_cell.alignment = copy(cell.alignment)
                dest_cell.border = copy(cell.border)
                dest_cell.number_format = cell.number_format
    for col_letter, dim in src_ws.column_dimensions.items():
        dest_ws.column_dimensions[col_letter].width = dim.width
    src_wb.close()


def generate_workbook(source_path, window_size, output_dir, base_config):
    """Generate a workbook with full calculation sheets for a given window."""
    print(f"\n{'='*60}")
    print(f"  Generating FULL workbook for {window_size}-month window")
    print(f"{'='*60}")

    master_df = load_recovery_triangle(source_path)
    n_total = len(master_df)
    balance_matrix = extract_balance_matrix(master_df)
    eads = master_df['EAD'].values.astype(float)
    master_tids = master_df['TID'].values.astype(float)

    config = ModelConfig(
        discount_rate=base_config.discount_rate,
        window_size=window_size,
        max_tid=base_config.max_tid,
        lgd_cap=base_config.lgd_cap,
        ci_percentile=base_config.ci_percentile,
    )

    # Run scenario to get aligned vintage results + backtest
    scenario = run_scenario(master_df, window_size, base_config, store_detail=False)
    if scenario is None:
        print(f"  SKIPPED: insufficient data")
        return None

    vr_list = scenario.vintage_results
    bt = scenario.backtest
    n_v = len(vr_list)
    n_tid = config.max_tid + 1

    print(f"  Vintages: {n_v}, Residuals: {scenario.n_residuals}")
    print(f"  RMSE: {scenario.rmse:.4f}, MAE: {scenario.mae:.4f}, Bias: {scenario.mean_error:+.4f}")

    wb = Workbook()
    wb.remove(wb.active)

    # 1. Copy master data
    copy_master_sheet(source_path, wb)

    # 2. LGD Term Structure Summary
    ws_sum = wb.create_sheet('LGD Term Structure Summary')
    _write_val(ws_sum, 1, 1, 'Vintage', font=HDR_FONT); ws_sum.cell(1, 1).fill = HDR_FILL
    _write_val(ws_sum, 1, 2, 'Period', font=HDR_FONT); ws_sum.cell(1, 2).fill = HDR_FILL
    for t in range(n_tid):
        cell = ws_sum.cell(row=1, column=3 + t, value=t)
        _style_hdr(cell)
    for i, v in enumerate(vr_list):
        r = 2 + i
        _write_val(ws_sum, r, 1, v.vintage_label)
        period_val = v.period
        if hasattr(period_val, 'to_pydatetime'):
            period_val = period_val.to_pydatetime()
        _write_val(ws_sum, r, 2, period_val, fmt=DATE_FMT)
        for t in range(min(n_tid, len(v.lgd_term_structure))):
            _write_val(ws_sum, r, 3 + t, float(v.lgd_term_structure[t]), fmt=PCT_FMT)
    ws_sum.column_dimensions['A'].width = 18
    ws_sum.column_dimensions['B'].width = 12

    # 3. Individual vintage calculation sheets (FULL audit trail)
    for vi, v in enumerate(vr_list):
        # Reconstruct the masked balance window for this vintage
        start_idx = v.start_idx
        end_idx = v.end_idx
        # Use full TID range (capped at max_tid, NOT window_size).
        # window_size controls how many COHORTS (rows), not TID columns.
        # The observation mask handles data availability per cohort.
        n_periods = min(config.max_tid, balance_matrix.shape[1])

        window = balance_matrix[start_idx:end_idx, :n_periods].copy()
        eads_window = eads[start_idx:end_idx].copy()
        tids_window = master_tids[start_idx:end_idx].copy()

        # Apply the observation mask
        offset = n_total - end_idx
        for i in range(window.shape[0]):
            cohort_master_idx = start_idx + i
            adjusted_max_tid = int(master_tids[cohort_master_idx]) - offset
            if adjusted_max_tid < n_periods:
                window[i, max(0, adjusted_max_tid + 1):] = np.nan

        # Compute intermediates
        recoveries = compute_aggregate_recoveries(window)
        cum_bal = compute_cumulative_balances(window)
        discount_mat = compute_discount_matrix(config.discount_rate, n_periods)
        lgd_ts = compute_lgd_term_structure(recoveries, cum_bal, discount_mat, cap=config.lgd_cap)

        # Weighted LGD
        cohort_lgds = np.full(window.shape[0], np.nan)
        for ci in range(window.shape[0]):
            row_valid = ~np.isnan(window[ci, :])
            if row_valid.any():
                first_valid = np.where(row_valid)[0][0]
                if first_valid < len(lgd_ts):
                    cohort_lgds[ci] = lgd_ts[first_valid]
        weighted_lgd = compute_ead_weighted_lgd(cohort_lgds, eads_window)

        # Cohort indices
        cohort_indices = np.arange(start_idx, end_idx)

        # Pad lgd_ts to max_tid+1 with 1.0
        if len(lgd_ts) < n_tid:
            lgd_ts = np.concatenate([lgd_ts, np.ones(n_tid - len(lgd_ts))])

        write_vintage_calc_sheet(
            wb, v.vintage_label, window, eads_window, tids_window,
            n_periods, config, recoveries, cum_bal, discount_mat,
            lgd_ts, weighted_lgd, cohort_indices,
        )

    # 4. LGD Backtest Summary
    ws_bt = wb.create_sheet('LGD Backtest Summary')
    vintage_labels = bt.vintage_labels
    periods = bt.periods
    hindsights = list(range(n_v, 0, -1))

    row = write_backtest_section(ws_bt, 1, 'FORECAST LGD BY VINTAGE',
                                 vintage_labels, periods, hindsights,
                                 bt.forecast_matrix, n_tid)
    row = write_backtest_section(ws_bt, row, 'ACTUAL LGD BY VINTAGE (DIAGONAL)',
                                 vintage_labels, periods, hindsights,
                                 bt.actual_matrix, n_tid)

    # Mean matrix
    mean_matrix = np.full_like(bt.forecast_matrix, np.nan)
    for i in range(n_v):
        for t in range(n_tid):
            if not np.isnan(bt.actual_matrix[i, t]):
                mean_matrix[i, t] = bt.mean_lgd[t]
    row = write_backtest_section(ws_bt, row, 'MEAN LGD BY VINTAGE',
                                 vintage_labels, periods, hindsights,
                                 mean_matrix, n_tid)

    # Upper CI
    z_score = bt.z_score
    ci_pctl = bt.ci_percentile
    ws_bt.cell(row=row, column=3, value=z_score).font = DATA_FONT
    ws_bt.cell(row=row, column=5, value=1).font = DATA_FONT
    row = write_backtest_section(ws_bt, row,
                                 f'UPPER BOUND ({ci_pctl*100:.0f}th pctl) \u2014 Binomial CI',
                                 vintage_labels, periods, hindsights,
                                 bt.upper_ci, n_tid)

    # Lower CI
    ws_bt.cell(row=row, column=3, value=-z_score).font = DATA_FONT
    row = write_backtest_section(ws_bt, row,
                                 f'LOWER BOUND ({(1-ci_pctl)*100:.0f}th pctl) \u2014 Binomial CI',
                                 vintage_labels, periods, hindsights,
                                 bt.lower_ci, n_tid)

    # Residuals
    row = write_backtest_section(ws_bt, row, 'DIFFERENCE (ACTUAL - FORECAST)',
                                 vintage_labels, periods, hindsights,
                                 bt.residual_matrix, n_tid)

    # Summary stats
    row += 1
    ws_bt.cell(row=row, column=1, value='BACKTEST SUMMARY STATISTICS').font = SECTION_FONT
    row += 1
    stats = [
        ('Window Size (months)', window_size),
        ('Number of Vintages', n_v),
        ('Number of Residuals', scenario.n_residuals),
        ('Discount Rate', base_config.discount_rate),
        ('CI Percentile', ci_pctl),
        ('z-score', z_score),
        ('', ''),
        ('Mean Error (Bias)', scenario.mean_error),
        ('Std Dev of Errors', scenario.std_error),
        ('RMSE', scenario.rmse),
        ('MAE', scenario.mae),
        ('Median AE', scenario.median_ae),
        ('Max |Error|', scenario.max_abs_error),
        ('IQR', scenario.iqr_error),
        ('Composite Score', scenario.composite_score),
        ('', ''),
        ('JB Rejects Normality', scenario.jb_reject),
        ('Overall CI Coverage', bt.overall_coverage),
        ('Latest LGD at TID=0', scenario.latest_lgd_tid0),
        ('Latest Weighted LGD', scenario.latest_weighted_lgd),
    ]
    for label, val in stats:
        _write_val(ws_bt, row, 1, label, font=LABEL_FONT)
        cell = _write_val(ws_bt, row, 2, val)
        if isinstance(val, float) and abs(val) < 2:
            cell.number_format = '0.000000'
        row += 1

    ws_bt.column_dimensions['A'].width = 22
    ws_bt.column_dimensions['B'].width = 14
    ws_bt.column_dimensions['C'].width = 12

    # Save
    filename = f'Munic_dashboard_LPU_1_{window_size}m.xlsx'
    output_path = os.path.join(output_dir, filename)
    wb.save(output_path)
    print(f"  Saved: {output_path}")
    return output_path


def main():
    source_path = os.path.join(os.path.dirname(__file__), '..', 'Munic_dashboard_LPU_1.xlsx')
    if not os.path.exists(source_path):
        source_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'Munic_dashboard_LPU_1.xlsx')

    output_dir = os.path.dirname(os.path.abspath(source_path))

    base_config = ModelConfig(
        discount_rate=0.15,
        window_size=60,
        max_tid=60,
        lgd_cap=None,
        ci_percentile=0.75,
    )

    window_sizes = [12, 18, 24, 30, 36, 42, 48, 54]

    print("=" * 60)
    print("  GENERATING PER-WINDOW WORKBOOKS (FULL CALCULATION SHEETS)")
    print("=" * 60)
    print(f"  Source: {source_path}")
    print(f"  Output: {output_dir}")
    print(f"  Windows: {window_sizes}")

    generated = []
    for ws_size in window_sizes:
        path = generate_workbook(source_path, ws_size, output_dir, base_config)
        if path:
            generated.append(path)

    print(f"\n{'='*60}")
    print(f"  COMPLETE: {len(generated)} workbooks generated")
    print(f"{'='*60}")
    for p in generated:
        print(f"  - {os.path.basename(p)}")


if __name__ == '__main__':
    main()
