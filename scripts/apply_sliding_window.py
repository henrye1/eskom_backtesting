"""Apply sliding window formula modifications to base workbook.

Replicates the user's 12m workbook formula changes for any window size W.
The key change: instead of using ALL cohorts in recovery and cumbal formulas,
use only the last W cohorts that have data at each TID column, creating a
sliding window that shifts backward for higher TIDs.

Pattern for TID column t (0-indexed):
  end_row = max(W + 1, last_data_row - t)
  start_row = end_row - W + 1
  Floor: if start_row < 2, set start_row = 2, end_row = W + 1
"""

import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula


def compute_row_range(tid_col: int, window_size: int, last_data_row: int = 61) -> tuple[int, int]:
    """Compute start/end rows for a given TID column and window size."""
    end_row = max(window_size + 1, last_data_row - tid_col)
    start_row = end_row - window_size + 1
    if start_row < 2:
        start_row = 2
        end_row = window_size + 1
    return start_row, end_row


def modify_vintage_sheet(ws, window_size: int, n_cohorts: int = 60, n_tid: int = 60):
    """Modify recovery and cumbal formulas in a single vintage calculation sheet."""
    last_data_row = n_cohorts + 1  # Row 61 for 60 cohorts (rows 2-61)
    recovery_row = 71
    cumbal_start_row = 73

    # Modify recovery row (row 71)
    for t in range(n_tid):
        col = 8 + t  # Column H = 8 for TID 0
        next_col = col + 1
        col_letter = get_column_letter(col)
        next_col_letter = get_column_letter(next_col)

        start_row, end_row = compute_row_range(t, window_size, last_data_row)

        formula_text = (
            f"=SUM(IF({next_col_letter}${start_row}:{next_col_letter}${end_row}<>\"\","
            f"{col_letter}${start_row}:{col_letter}${end_row}))"
            f"-SUM({next_col_letter}${start_row}:{next_col_letter}${end_row})"
        )

        cell_ref = f"{col_letter}{recovery_row}"
        ws[cell_ref] = ArrayFormula(ref=cell_ref, text=formula_text)

    # Modify cumbal rows (rows 73 to 73 + n_tid - 1)
    for r in range(n_tid):
        cumbal_row = cumbal_start_row + r
        bal_col = 8 + r  # Balance column for this TID row
        bal_col_letter = get_column_letter(bal_col)

        for c in range(r, n_tid):
            col = 8 + c
            cond_col = col + 1  # Condition: check data at c+1
            cond_col_letter = get_column_letter(cond_col)
            col_letter = get_column_letter(col)

            start_row, end_row = compute_row_range(c, window_size, last_data_row)

            formula_text = (
                f"=SUM(IF({cond_col_letter}${start_row}:{cond_col_letter}${end_row}<>\"\","
                f"${bal_col_letter}${start_row}:${bal_col_letter}${end_row}))"
            )

            cell_ref = f"{col_letter}{cumbal_row}"
            ws[cell_ref] = ArrayFormula(ref=cell_ref, text=formula_text)


def apply_sliding_window(base_path: str, output_path: str, window_size: int):
    """Copy base workbook and apply sliding window formulas."""
    shutil.copy2(base_path, output_path)

    wb = openpyxl.load_workbook(output_path)

    vintage_sheets = [s for s in wb.sheetnames if s.startswith('RR LGD TERM STRUCTURES  (')]
    print(f"Processing {len(vintage_sheets)} vintage sheets for {window_size}m window...")

    for sheet_name in vintage_sheets:
        ws = wb[sheet_name]
        modify_vintage_sheet(ws, window_size)

    wb.save(output_path)
    print(f"Saved: {output_path}")


def main():
    base_path = Path(__file__).parent.parent / "Munic_dashboard_LPU_1.xlsx"
    output_dir = Path(__file__).parent.parent

    window_sizes = [18, 24, 30, 36, 42, 48, 54]

    if len(sys.argv) > 1:
        window_sizes = [int(w) for w in sys.argv[1].split(',')]

    for w in window_sizes:
        output_path = output_dir / f"Munic_dashboard_LPU_1_{w}m.xlsx"
        apply_sliding_window(str(base_path), str(output_path), w)

    print(f"\nDone! Generated {len(window_sizes)} workbooks.")


if __name__ == "__main__":
    main()
