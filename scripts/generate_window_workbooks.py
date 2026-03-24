"""Generate per-window-size workbooks mirroring the Munic_dashboard_LPU_1 structure.

For each window size (12, 18, 24, 30, 36, 42, 48, 54), produces a workbook with:
  - RR LGD TERM STRUCTURE ALL (master data, unchanged)
  - LGD Term Structure Summary (vintage LGD curves)
  - LGD Backtest Summary (forecast, actual, mean, upper CI, lower CI, residuals)
  - Individual vintage sheets (RR LGD TERM STRUCTURES (offset-offset))
"""

import os
import sys
import warnings
import datetime

import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from scipy.stats import norm

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from lgd_model.config import ModelConfig
from lgd_model.data_loader import load_recovery_triangle, extract_balance_matrix
from lgd_model.vintage import run_vintage_analysis
from lgd_model.backtest import run_backtest
from lgd_model.scenario import run_scenario


# Formatting constants
HEADER_FILL = PatternFill('solid', fgColor='2E75B6')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
SECTION_FONT = Font(name='Arial', bold=True, size=11, color='2E75B6')
DATA_FONT = Font(name='Arial', size=9)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
PCT_FMT = '0.0000%'
NUM_FMT = '0.0000'
DATE_FMT = 'YYYY-MM-DD'


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(cell, is_pct=True):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    if is_pct and isinstance(cell.value, (int, float)) and not np.isnan(cell.value):
        cell.number_format = PCT_FMT
    cell.alignment = Alignment(horizontal='center')


def write_matrix_section(ws, start_row, title, header_labels, vintage_labels, periods,
                         hindsights, matrix, n_tid, tid_offset=0):
    """Write a labelled section (forecast, actual, mean, CI, residuals) to a sheet."""
    # Section title
    ws.cell(row=start_row, column=1, value=title).font = SECTION_FONT

    # Header row
    hdr_row = start_row + 1
    ws.cell(row=hdr_row, column=1, value='Vintage')
    ws.cell(row=hdr_row, column=2, value='Period')
    ws.cell(row=hdr_row, column=3, value='Hindsight')
    for t in range(n_tid):
        ws.cell(row=hdr_row, column=4 + t, value=t)
    style_header_row(ws, hdr_row, 3 + n_tid)

    # Data rows
    n_v = len(vintage_labels)
    for i in range(n_v):
        r = hdr_row + 1 + i
        ws.cell(row=r, column=1, value=vintage_labels[i]).font = DATA_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER

        period_val = periods[i]
        if hasattr(period_val, 'to_pydatetime'):
            period_val = period_val.to_pydatetime()
        ws.cell(row=r, column=2, value=period_val).font = DATA_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=2).number_format = DATE_FMT

        ws.cell(row=r, column=3, value=hindsights[i]).font = DATA_FONT
        ws.cell(row=r, column=3).border = THIN_BORDER

        for t in range(min(n_tid, matrix.shape[1])):
            val = matrix[i, t]
            cell = ws.cell(row=r, column=4 + t)
            if np.isnan(val):
                cell.value = None
            else:
                cell.value = float(val)
            style_data_cell(cell, is_pct=True)

    return hdr_row + 1 + n_v + 1  # next available row (with blank line)


def copy_master_sheet(source_path, dest_wb):
    """Copy the RR LGD TERM STRUCTURE ALL sheet from source to dest."""
    src_wb = load_workbook(source_path, data_only=True)
    src_ws = src_wb['RR LGD TERM STRUCTURE ALL']
    dest_ws = dest_wb.create_sheet('RR LGD TERM STRUCTURE ALL')

    for row in src_ws.iter_rows():
        for cell in row:
            dest_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dest_cell.font = cell.font.copy()
                dest_cell.fill = cell.fill.copy()
                dest_cell.alignment = cell.alignment.copy()
                dest_cell.border = cell.border.copy()
                dest_cell.number_format = cell.number_format

    # Copy column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        dest_ws.column_dimensions[col_letter].width = dim.width

    src_wb.close()
    return dest_ws


def generate_workbook_for_window(source_path, window_size, output_dir, base_config):
    """Generate a complete workbook for a given window size."""
    print(f"\n{'='*60}")
    print(f"  Generating workbook for {window_size}-month window")
    print(f"{'='*60}")

    # Run the model
    master_df = load_recovery_triangle(source_path)
    n_total = len(master_df)

    config = ModelConfig(
        discount_rate=base_config.discount_rate,
        window_size=window_size,
        max_tid=base_config.max_tid,
        lgd_cap=base_config.lgd_cap,
        ci_percentile=base_config.ci_percentile,
    )

    # Run scenario to get aligned vintages
    scenario = run_scenario(master_df, window_size, base_config, store_detail=False)
    if scenario is None:
        print(f"  SKIPPED: insufficient data for {window_size}m window")
        return None

    vr_list = scenario.vintage_results
    bt = scenario.backtest
    n_v = len(vr_list)
    n_tid = config.max_tid + 1

    print(f"  Vintages: {n_v}")
    print(f"  Residuals: {scenario.n_residuals}")
    print(f"  RMSE: {scenario.rmse:.4f}, MAE: {scenario.mae:.4f}, Bias: {scenario.mean_error:+.4f}")

    # Create workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # 1. Copy master data sheet
    copy_master_sheet(source_path, wb)

    # 2. LGD Term Structure Summary
    ws_summary = wb.create_sheet('LGD Term Structure Summary')
    ws_summary.cell(row=1, column=1, value='Vintage')
    ws_summary.cell(row=1, column=2, value='Period')
    for t in range(n_tid):
        ws_summary.cell(row=1, column=3 + t, value=t)
    style_header_row(ws_summary, 1, 2 + n_tid)

    for i, v in enumerate(vr_list):
        r = 2 + i
        ws_summary.cell(row=r, column=1, value=v.vintage_label).font = DATA_FONT
        ws_summary.cell(row=r, column=1).border = THIN_BORDER

        period_val = v.period
        if hasattr(period_val, 'to_pydatetime'):
            period_val = period_val.to_pydatetime()
        ws_summary.cell(row=r, column=2, value=period_val).font = DATA_FONT
        ws_summary.cell(row=r, column=2).border = THIN_BORDER
        ws_summary.cell(row=r, column=2).number_format = DATE_FMT

        for t in range(min(n_tid, len(v.lgd_term_structure))):
            cell = ws_summary.cell(row=r, column=3 + t, value=float(v.lgd_term_structure[t]))
            style_data_cell(cell)

    ws_summary.column_dimensions['A'].width = 18
    ws_summary.column_dimensions['B'].width = 12

    # 3. LGD Backtest Summary
    ws_bt = wb.create_sheet('LGD Backtest Summary')

    vintage_labels = bt.vintage_labels
    periods = bt.periods
    hindsights = list(range(n_v, 0, -1))  # n_v for oldest, 1 for newest

    # Forecast section
    row = write_matrix_section(
        ws_bt, 1, 'FORECAST LGD BY VINTAGE',
        None, vintage_labels, periods, hindsights,
        bt.forecast_matrix, n_tid
    )

    # Actual section
    row = write_matrix_section(
        ws_bt, row, 'ACTUAL LGD BY VINTAGE (DIAGONAL)',
        None, vintage_labels, periods, hindsights,
        bt.actual_matrix, n_tid
    )

    # Mean section
    mean_matrix = np.full_like(bt.forecast_matrix, np.nan)
    for i in range(n_v):
        for t in range(n_tid):
            if not np.isnan(bt.actual_matrix[i, t]):
                mean_matrix[i, t] = bt.mean_lgd[t]

    row = write_matrix_section(
        ws_bt, row, 'MEAN LGD BY VINTAGE',
        None, vintage_labels, periods, hindsights,
        mean_matrix, n_tid
    )

    # Upper CI section
    z_score = bt.z_score
    ci_pctl = bt.ci_percentile
    ws_bt.cell(row=row, column=1,
               value=f'UPPER BOUND ({ci_pctl*100:.0f}th pctl) — Binomial CI').font = SECTION_FONT
    ws_bt.cell(row=row, column=3, value=z_score).font = DATA_FONT
    row_upper_start = row
    row = write_matrix_section(
        ws_bt, row, f'UPPER BOUND ({ci_pctl*100:.0f}th pctl) — Binomial CI',
        None, vintage_labels, periods, hindsights,
        bt.upper_ci, n_tid
    )
    # Remove the duplicate title cell
    ws_bt.cell(row=row_upper_start, column=1, value=None)

    # Lower CI section
    ws_bt.cell(row=row, column=1,
               value=f'LOWER BOUND ({(1-ci_pctl)*100:.0f}th pctl) — Binomial CI').font = SECTION_FONT
    ws_bt.cell(row=row, column=3, value=-z_score).font = DATA_FONT
    row_lower_start = row
    row = write_matrix_section(
        ws_bt, row, f'LOWER BOUND ({(1-ci_pctl)*100:.0f}th pctl) — Binomial CI',
        None, vintage_labels, periods, hindsights,
        bt.lower_ci, n_tid
    )
    ws_bt.cell(row=row_lower_start, column=1, value=None)

    # Residuals section
    row = write_matrix_section(
        ws_bt, row, 'DIFFERENCE (ACTUAL - FORECAST)',
        None, vintage_labels, periods, hindsights,
        bt.residual_matrix, n_tid
    )

    # Summary statistics section
    row += 1
    ws_bt.cell(row=row, column=1, value='BACKTEST SUMMARY STATISTICS').font = SECTION_FONT
    row += 1

    stats_items = [
        ('Window Size (months)', window_size),
        ('Number of Vintages', n_v),
        ('Number of Residuals', scenario.n_residuals),
        ('Discount Rate', base_config.discount_rate),
        ('CI Percentile', ci_pctl),
        ('z-score', z_score),
        ('', ''),
        ('Mean Error (Bias)', scenario.mean_error),
        ('Std Dev of Errors', scenario.std_error),
        ('RMSE', scenario.rmse),
        ('MAE', scenario.mae),
        ('Median AE', scenario.median_ae),
        ('Max |Error|', scenario.max_abs_error),
        ('IQR', scenario.iqr_error),
        ('Composite Score', scenario.composite_score),
        ('AIC Proxy', scenario.aic_proxy),
        ('', ''),
        ('Jarque-Bera Rejects Normality', scenario.jb_reject),
        ('N (residuals)', bt.normality_stats.get('n', '')),
        ('Skewness', bt.normality_stats.get('skewness', '')),
        ('Excess Kurtosis', bt.normality_stats.get('excess_kurtosis', '')),
        ('JB Statistic', bt.normality_stats.get('jarque_bera_stat', '')),
        ('JB Critical (5%)', bt.normality_stats.get('jb_critical_005', '')),
        ('Chi-Sq Statistic', bt.normality_stats.get('chi_sq_stat', '')),
        ('Chi-Sq Critical (5%)', bt.normality_stats.get('chi_sq_critical_005', '')),
        ('', ''),
        ('Overall CI Coverage', bt.overall_coverage),
        ('Latest LGD at TID=0', scenario.latest_lgd_tid0),
        ('Latest Weighted LGD', scenario.latest_weighted_lgd),
    ]

    for label, val in stats_items:
        ws_bt.cell(row=row, column=1, value=label).font = Font(name='Arial', bold=True, size=9)
        ws_bt.cell(row=row, column=1).border = THIN_BORDER
        cell = ws_bt.cell(row=row, column=2, value=val)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if isinstance(val, float) and abs(val) < 2:
            cell.number_format = '0.000000'
        row += 1

    ws_bt.column_dimensions['A'].width = 22
    ws_bt.column_dimensions['B'].width = 14
    ws_bt.column_dimensions['C'].width = 12

    # 4. Individual vintage sheets
    for i, v in enumerate(vr_list):
        sheet_name = f'RR LGD TERM STRUCTURES  ({v.vintage_label.replace("Latest ", "")})'
        # Truncate sheet name if too long (Excel limit = 31 chars)
        if len(sheet_name) > 31:
            label_part = v.vintage_label.replace("Latest ", "")
            sheet_name = f'TERM STRUCT ({label_part})'
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

        ws_v = wb.create_sheet(sheet_name)

        # Vintage metadata
        ws_v.cell(row=1, column=1, value='Vintage').font = Font(name='Arial', bold=True, size=9)
        ws_v.cell(row=1, column=2, value=v.vintage_label).font = DATA_FONT
        ws_v.cell(row=2, column=1, value='Window').font = Font(name='Arial', bold=True, size=9)
        ws_v.cell(row=2, column=2, value=f'{window_size} months').font = DATA_FONT
        ws_v.cell(row=3, column=1, value='Cohorts').font = Font(name='Arial', bold=True, size=9)
        ws_v.cell(row=3, column=2, value=v.n_cohorts).font = DATA_FONT
        ws_v.cell(row=4, column=1, value='Period').font = Font(name='Arial', bold=True, size=9)
        period_val = v.period
        if hasattr(period_val, 'to_pydatetime'):
            period_val = period_val.to_pydatetime()
        ws_v.cell(row=4, column=2, value=period_val).font = DATA_FONT
        ws_v.cell(row=4, column=2).number_format = DATE_FMT
        ws_v.cell(row=5, column=1, value='Weighted LGD').font = Font(name='Arial', bold=True, size=9)
        ws_v.cell(row=5, column=2, value=float(v.weighted_lgd)).font = DATA_FONT
        ws_v.cell(row=5, column=2).number_format = PCT_FMT

        # LGD Term Structure
        ws_v.cell(row=7, column=1, value='LGD TERM STRUCTURE').font = SECTION_FONT
        ws_v.cell(row=8, column=1, value='TID')
        ws_v.cell(row=8, column=2, value='LGD')
        ws_v.cell(row=8, column=3, value='Recovery Rate')
        style_header_row(ws_v, 8, 3)

        for t in range(min(n_tid, len(v.lgd_term_structure))):
            r = 9 + t
            ws_v.cell(row=r, column=1, value=t).font = DATA_FONT
            ws_v.cell(row=r, column=1).border = THIN_BORDER
            lgd_val = float(v.lgd_term_structure[t])
            cell_lgd = ws_v.cell(row=r, column=2, value=lgd_val)
            style_data_cell(cell_lgd)
            cell_rr = ws_v.cell(row=r, column=3, value=1.0 - lgd_val)
            style_data_cell(cell_rr)

        ws_v.column_dimensions['A'].width = 14
        ws_v.column_dimensions['B'].width = 14
        ws_v.column_dimensions['C'].width = 14

    # Save
    filename = f'Munic_dashboard_LPU_1_{window_size}m.xlsx'
    output_path = os.path.join(output_dir, filename)
    wb.save(output_path)
    print(f"  Saved: {output_path}")
    return output_path


def main():
    source_path = os.path.join(os.path.dirname(__file__), '..', 'Munic_dashboard_LPU_1.xlsx')
    if not os.path.exists(source_path):
        source_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'Munic_dashboard_LPU_1.xlsx')

    output_dir = os.path.dirname(os.path.abspath(source_path))

    base_config = ModelConfig(
        discount_rate=0.15,
        window_size=60,
        max_tid=60,
        lgd_cap=None,
        ci_percentile=0.75,
    )

    window_sizes = [12, 18, 24, 30, 36, 42, 48, 54]

    print("=" * 60)
    print("  GENERATING PER-WINDOW WORKBOOKS")
    print("=" * 60)
    print(f"  Source: {source_path}")
    print(f"  Output: {output_dir}")
    print(f"  Windows: {window_sizes}")
    print(f"  Discount rate: {base_config.discount_rate:.1%}")
    print(f"  CI percentile: {base_config.ci_percentile:.0%}")

    generated = []
    for ws in window_sizes:
        path = generate_workbook_for_window(source_path, ws, output_dir, base_config)
        if path:
            generated.append(path)

    print(f"\n{'='*60}")
    print(f"  COMPLETE: {len(generated)} workbooks generated")
    print(f"{'='*60}")
    for p in generated:
        print(f"  - {os.path.basename(p)}")


if __name__ == '__main__':
    main()
