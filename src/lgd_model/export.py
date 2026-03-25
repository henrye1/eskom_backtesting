"""Excel export functions for LGD model results.

Includes three export levels:
1. Summary exports (export_results_to_excel, export_multi_scenario_excel)
2. Full audit trail workbooks (export_full_audit_workbook) — replicates
   the reference workbook structure with LIVE EXCEL FORMULAS for all
   intermediate calculations so auditors can trace every result.
"""

import io
import zipfile

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from lgd_model.backtest import BacktestResult
from lgd_model.config import ModelConfig
from lgd_model.data_loader import extract_balance_matrix
from lgd_model.scenario import ScenarioResult, generate_scenario_comparison_table
from lgd_model.vintage import VintageResult

# ── Formatting constants for audit trail workbooks ───────────────────
_HDR_FILL = PatternFill('solid', fgColor='2E75B6')
_HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=9)
_SECTION_FONT = Font(name='Arial', bold=True, size=10, color='2E75B6')
_LABEL_FONT = Font(name='Arial', bold=True, size=9)
_DATA_FONT = Font(name='Arial', size=9)
_PARAM_FONT = Font(name='Arial', bold=True, size=9, color='333333')
_THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
_PCT_FMT = '0.0000%'
_NUM_FMT = '#,##0.00'
_DATE_FMT = 'YYYY-MM-DD'


# ── Cell reference helpers for formula generation ────────────────────

def _ref(row: int, col: int) -> str:
    """1-based (row, col) to cell reference, e.g. 'H2'."""
    return f"{get_column_letter(col)}{row}"


def _abs_ref(row: int, col: int) -> str:
    """Absolute reference, e.g. '$H$2'."""
    return f"${get_column_letter(col)}${row}"


def _col_rng(col: int, r1: int, r2: int) -> str:
    """Column range, e.g. 'H2:H61'."""
    return f"{get_column_letter(col)}{r1}:{get_column_letter(col)}{r2}"


def _row_rng(row: int, c1: int, c2: int) -> str:
    """Row range, e.g. 'H5:BM5'."""
    return f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}"


# ── Styling helpers ──────────────────────────────────────────────────

def _style_hdr(cell):
    cell.font = _HDR_FONT
    cell.fill = _HDR_FILL
    cell.alignment = Alignment(horizontal='center')
    cell.border = _THIN_BORDER


def _style_data(cell, fmt=None):
    cell.font = _DATA_FONT
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(horizontal='center')
    if fmt:
        cell.number_format = fmt


def _hide_gridlines(ws):
    """Hide gridlines on a worksheet."""
    ws.sheet_view.showGridLines = False


def _write_val(ws, row, col, val, fmt=None, font=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = font or _DATA_FONT
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(horizontal='center')
    if fmt:
        cell.number_format = fmt
    return cell


# ═════════════════════════════════════════════════════════════════════
# Summary Exports (values only — quick reference)
# ═════════════════════════════════════════════════════════════════════

def generate_summary_dataframe(
    vintage_results: list[VintageResult],
    config: ModelConfig,
) -> pd.DataFrame:
    """Build a summary DataFrame of LGD term structures by vintage."""
    rows = []
    for vr in vintage_results:
        row: dict = {'Vintage': vr.vintage_label, 'Period': vr.period}
        for t in range(config.max_tid + 1):
            row[t] = (
                vr.lgd_term_structure[t]
                if t < len(vr.lgd_term_structure)
                else np.nan
            )
        rows.append(row)
    return pd.DataFrame(rows)


def export_results_to_excel(
    vintage_results: list[VintageResult],
    bt: BacktestResult,
    config: ModelConfig,
    output_path: str,
) -> None:
    """Export single-window results to an Excel workbook."""
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary_df = generate_summary_dataframe(vintage_results, config)
        summary_df.to_excel(writer, sheet_name='LGD Term Structure Summary', index=False)

        forecast_df = pd.DataFrame(
            bt.forecast_matrix,
            columns=[f'TID_{t}' for t in range(bt.forecast_matrix.shape[1])],
            index=bt.vintage_labels,
        )
        forecast_df.insert(0, 'Period', bt.periods)
        forecast_df.to_excel(writer, sheet_name='Forecast LGD')

        actual_df = pd.DataFrame(
            bt.actual_matrix,
            columns=[f'TID_{t}' for t in range(bt.actual_matrix.shape[1])],
            index=bt.vintage_labels,
        )
        actual_df.insert(0, 'Period', bt.periods)
        actual_df.to_excel(writer, sheet_name='Actual LGD')

        resid_df = pd.DataFrame(
            bt.residual_matrix,
            columns=[f'TID_{t}' for t in range(bt.residual_matrix.shape[1])],
            index=bt.vintage_labels,
        )
        resid_df.insert(0, 'Period', bt.periods)
        resid_df.to_excel(writer, sheet_name='Residuals')

        stats_data = {
            'Metric': [
                'N', 'Mean', 'Std Dev', 'Skewness', 'Excess Kurtosis',
                'JB Stat', 'JB Critical', 'JB Reject',
                'Chi-Sq Stat', 'Chi-Sq Critical', 'Chi-Sq Reject', 'CI Method',
            ],
            'Value': [
                bt.normality_stats['n'],
                bt.normality_stats['mean'],
                bt.normality_stats['std'],
                bt.normality_stats['skewness'],
                bt.normality_stats['excess_kurtosis'],
                bt.normality_stats['jarque_bera_stat'],
                bt.normality_stats['jb_critical_005'],
                bt.normality_stats['jb_reject'],
                bt.normality_stats.get('chi_sq_stat', np.nan),
                bt.normality_stats.get('chi_sq_critical_005', np.nan),
                bt.normality_stats.get('chi_sq_reject', None),
                'sem_scaled',
            ],
        }
        pd.DataFrame(stats_data).to_excel(
            writer, sheet_name='Normality Tests', index=False
        )

        upper_df = pd.DataFrame(
            bt.upper_ci,
            columns=[f'TID_{t}' for t in range(bt.upper_ci.shape[1])],
            index=bt.vintage_labels,
        )
        upper_df.insert(0, 'Period', bt.periods)
        upper_df.to_excel(writer, sheet_name='Upper CI (SEM Scaled)')

        lower_df = pd.DataFrame(
            bt.lower_ci,
            columns=[f'TID_{t}' for t in range(bt.lower_ci.shape[1])],
            index=bt.vintage_labels,
        )
        lower_df.insert(0, 'Period', bt.periods)
        lower_df.to_excel(writer, sheet_name='Lower CI (SEM Scaled)')

        # Hide gridlines on all sheets
        for ws in writer.book.worksheets:
            _hide_gridlines(ws)

    print(f"\n  Results exported to: {output_path}")


def export_multi_scenario_excel(
    scenarios: list[ScenarioResult],
    output_path: str = 'LGD_Multi_Scenario_Output.xlsx',
) -> None:
    """Export all scenario results to a comprehensive Excel workbook."""
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        comp_df = generate_scenario_comparison_table(scenarios)
        comp_df.to_excel(writer, sheet_name='Scenario Comparison', index=False)

        for s in scenarios:
            ws = s.window_size
            bt = s.backtest

            fc_df = pd.DataFrame(
                bt.forecast_matrix,
                columns=[f'TID_{t}' for t in range(bt.forecast_matrix.shape[1])],
                index=bt.vintage_labels,
            )
            fc_df.insert(0, 'Period', bt.periods)
            fc_df.to_excel(writer, sheet_name=f'{ws}m Forecast')

            res_df = pd.DataFrame(
                bt.residual_matrix,
                columns=[f'TID_{t}' for t in range(bt.residual_matrix.shape[1])],
                index=bt.vintage_labels,
            )
            res_df.insert(0, 'Period', bt.periods)
            res_df.to_excel(writer, sheet_name=f'{ws}m Residuals')

            latest = s.vintage_results[-1]
            lgd_df = pd.DataFrame({
                'TID': range(len(latest.lgd_term_structure)),
                'LGD': latest.lgd_term_structure,
                'Recovery': 1.0 - latest.lgd_term_structure,
            })
            lgd_df.to_excel(writer, sheet_name=f'{ws}m LGD Term', index=False)

        # Hide gridlines on all sheets
        for sheet in writer.book.worksheets:
            _hide_gridlines(sheet)

    print(f"\n  Multi-scenario results exported to: {output_path}")


# ═════════════════════════════════════════════════════════════════════
# Full Audit Trail Workbook — LIVE EXCEL FORMULAS
# ═════════════════════════════════════════════════════════════════════


def _compute_excel_cohort_range(
    tid_col: int, min_obs_window: int | None, n_cohorts: int, br1: int,
) -> tuple[int, int]:
    """Compute Excel row range for a given TID using sliding window logic.

    Mirrors _compute_cohort_range from core_engine but returns 1-based
    Excel row numbers (br1-based).
    """
    if min_obs_window is None:
        return br1, br1 + n_cohorts - 1
    end_idx = max(min_obs_window, n_cohorts - tid_col)
    start_idx = max(0, end_idx - min_obs_window)
    return br1 + start_idx, br1 + end_idx - 1


def _write_vintage_calc_sheet(
    wb, vintage_label, balance_window, eads_window,
    master_tids_window, n_periods, config, cohort_indices,
    weighted_lgd_value,
):
    """Write a vintage calculation sheet with LIVE Excel formulas.

    All intermediate calculations use formulas referencing the balance
    triangle so auditors can click any cell and trace the calculation:

    Balance Matrix (VALUES) → Recovery (FORMULA) → CumBal (FORMULA)
    → Discount (FORMULA) → LGD Components (FORMULA) → LGD Term Structure (FORMULA)
    → Weighted LGD (FORMULA)
    """
    sheet_label = vintage_label.replace("Latest ", "")
    sheet_name = f'RR LGD TERM STRUCTURES  ({sheet_label})'
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    ws = wb.create_sheet(sheet_name)
    n_cohorts = balance_window.shape[0]

    # Layout constants
    BC = 8           # Balance columns start at H (col 8)
    BR1 = 2          # First balance data row
    BR2 = n_cohorts + 1  # Last balance data row

    # ── Row 1: Header ──
    _write_val(ws, 1, 3, 'LGD TERM STRUCTURES', font=_SECTION_FONT)
    _write_val(ws, 1, 7, 'MIN', font=_LABEL_FONT)
    for t in range(n_periods + 1):
        _write_val(ws, 1, BC + t, t, font=_HDR_FONT)
        ws.cell(row=1, column=BC + t).fill = _HDR_FILL

    # ── Balance triangle (VALUES — raw data after observation mask) ──
    for i in range(n_cohorts):
        r = BR1 + i
        _write_val(ws, r, 1, int(cohort_indices[i]))
        _write_val(ws, r, 2, int(master_tids_window[i]))
        # Column C: per-cohort LGD — formula referencing Check row (written later)
        # (deferred to after Check row is positioned)

        # Column G: MIN balance — FORMULA
        bal_range = _row_rng(r, BC, BC + n_periods - 1)
        cell = ws.cell(row=r, column=7, value=f'=MIN({bal_range})')
        _style_data(cell, _NUM_FMT)

        # Balance values (raw data)
        for t in range(n_periods):
            val = balance_window[i, t]
            cell = ws.cell(row=r, column=BC + t)
            cell.value = None if np.isnan(val) else float(val)
            _style_data(cell, _NUM_FMT)

    # ── Check row + Weighted LGD ──
    CHECK_ROW = BR2 + 3
    _write_val(ws, CHECK_ROW, 1, 'Weighted', font=_LABEL_FONT)
    _write_val(ws, CHECK_ROW, 7, 'Check', font=_LABEL_FONT)

    # ── Parameters block ──
    PARAM_ROW = CHECK_ROW + 2
    _write_val(ws, PARAM_ROW, 7, 'Interest Rate', font=_PARAM_FONT)
    _write_val(ws, PARAM_ROW, 8, config.discount_rate)
    _write_val(ws, PARAM_ROW + 1, 7, 'Columns', font=_PARAM_FONT)
    _write_val(ws, PARAM_ROW + 1, 8, n_periods)
    _write_val(ws, PARAM_ROW + 2, 7, 'Rows', font=_PARAM_FONT)
    _write_val(ws, PARAM_ROW + 2, 8, n_cohorts)
    _write_val(ws, PARAM_ROW + 3, 7, 'Min', font=_PARAM_FONT)
    _write_val(ws, PARAM_ROW + 3, 8, 0)

    rate_ref = _abs_ref(PARAM_ROW, 8)

    # ── Recovery vector — FORMULAS ──
    # With min_obs_window: sliding cohort range per TID column
    # Reference 12m: TID 0 uses rows 50:61, TID 1 uses 49:60, etc.
    RECOV_ROW = PARAM_ROW + 5
    _write_val(ws, RECOV_ROW, 7, 'Recovery', font=_SECTION_FONT)
    mow = config.min_obs_window
    for n in range(n_periods):
        cell = ws.cell(row=RECOV_ROW, column=BC + n)
        if n < n_periods - 1:
            # Cohort range based on the CURRENT TID n (not n+1)
            r1, r2 = _compute_excel_cohort_range(n, mow, n_cohorts, BR1)
            cn = get_column_letter(BC + n)
            cn1 = get_column_letter(BC + n + 1)
            cell.value = (
                f'=SUMPRODUCT(({cn1}${r1}:{cn1}${r2}<>"")*'
                f'{cn}${r1}:{cn}${r2})'
                f'-SUM({cn1}${r1}:{cn1}${r2})'
            )
        else:
            cell.value = 0  # Last period recovery = 0
        _style_data(cell, _NUM_FMT)

    # ── Cumulative Balance matrix — FORMULAS ──
    # With min_obs_window: sliding cohort range per TID column
    CUMBAL_ROW = RECOV_ROW + 2
    for t_row in range(n_periods + 1):
        if t_row == 0:
            _write_val(ws, CUMBAL_ROW, 7, 0)
        else:
            ws.cell(row=CUMBAL_ROW + t_row, column=7,
                    value=f'={_ref(CUMBAL_ROW + t_row - 1, 7)}+1')
        for t_col in range(n_periods):
            cell = ws.cell(row=CUMBAL_ROW + t_row, column=BC + t_col)
            if t_col < t_row or t_row >= n_periods:
                cell.value = None
            else:
                # Condition column is always c+1
                cond_col_idx = t_col + 1
                r1, r2 = _compute_excel_cohort_range(t_col, mow, n_cohorts, BR1)
                br_letter = get_column_letter(BC + t_row)
                cond_letter = get_column_letter(BC + cond_col_idx)
                cell.value = (
                    f'=SUMPRODUCT(({cond_letter}${r1}:{cond_letter}${r2}<>"")*'
                    f'${br_letter}${r1}:${br_letter}${r2})'
                )
            _style_data(cell, _NUM_FMT)

    # ── Discount factor matrix — FORMULAS ──
    # Matches reference: =1/(1+$H$66)^((H$135-$G136)/12)
    # Header row starts at 1 (1-indexed), uses incremental formula
    DISC_HDR = CUMBAL_ROW + n_periods + 2
    DISC_ROW = DISC_HDR + 1
    _write_val(ws, DISC_HDR, BC, 1)  # Header starts at 1, not 0
    for t in range(1, n_periods + 1):
        ws.cell(row=DISC_HDR, column=BC + t,
                value=f'={_ref(DISC_HDR, BC + t - 1)}+1')
    for t_row in range(n_periods + 1):
        if t_row == 0:
            _write_val(ws, DISC_ROW, 7, 0)
        else:
            ws.cell(row=DISC_ROW + t_row, column=7,
                    value=f'={_ref(DISC_ROW + t_row - 1, 7)}+1')
        for t_col in range(n_periods):
            cell = ws.cell(row=DISC_ROW + t_row, column=BC + t_col)
            if t_col < t_row or t_row >= n_periods:
                cell.value = None
            else:
                # Reference header cell for column index, G cell for row index
                col_hdr = f'{get_column_letter(BC + t_col)}${DISC_HDR}'
                row_lbl = f'$G{DISC_ROW + t_row}'
                cell.value = f'=1/(1+{rate_ref})^(({col_hdr}-{row_lbl})/12)'
            _style_data(cell, '0.000000')

    # ── LGD component matrix — FORMULAS ──
    # Matches reference: =IFERROR(Recovery/CumBal, 0)
    # NOTE: NO discount multiplier here — discount applied in LGD formula
    COMP_ROW = DISC_ROW + n_periods + 3
    for t_row in range(n_periods + 1):
        if t_row == 0:
            _write_val(ws, COMP_ROW, 7, 0)
        else:
            ws.cell(row=COMP_ROW + t_row, column=7,
                    value=f'={_ref(COMP_ROW + t_row - 1, 7)}+1')
        for t_col in range(n_periods):
            cell = ws.cell(row=COMP_ROW + t_row, column=BC + t_col)
            if t_col < t_row or t_row >= n_periods:
                cell.value = None
            else:
                rec = f'{get_column_letter(BC + t_col)}${RECOV_ROW}'
                cb = _ref(CUMBAL_ROW + t_row, BC + t_col)
                cell.value = f'=IFERROR({rec}/{cb},0)'
            _style_data(cell, _PCT_FMT)

    # ── LGD term structure — FORMULAS (columns E/F starting at row 2) ──
    # Aligned with cohort rows so A2:A61 corresponds to E2:F62
    # Column E = TID index, Column F = LGD value
    LGD_R1 = BR1  # Start at row 2 to align with cohort data
    _write_val(ws, 1, 5, 'TID', font=_HDR_FONT)
    ws.cell(1, 5).fill = _HDR_FILL
    _write_val(ws, 1, 6, 'LGD', font=_HDR_FONT)
    ws.cell(1, 6).fill = _HDR_FILL
    cap = config.lgd_cap
    for t in range(n_periods):
        # Column E: TID index
        ws.cell(row=LGD_R1 + t, column=5, value=t)
        # Column F: LGD = IFERROR(1 - SUMPRODUCT(component_row, discount_row), 1)
        comp_rng = _row_rng(COMP_ROW + t, BC + t, BC + n_periods - 1)
        disc_rng = _row_rng(DISC_ROW + t, BC + t, BC + n_periods - 1)
        if cap is not None:
            inner = f'MIN({cap},1-SUMPRODUCT({comp_rng},{disc_rng}))'
        else:
            inner = f'1-SUMPRODUCT({comp_rng},{disc_rng})'
        formula = f'=IFERROR({inner},1)'
        cell = ws.cell(row=LGD_R1 + t, column=6, value=formula)
        _style_data(cell, _PCT_FMT)
    # Last TID = 1.0
    ws.cell(row=LGD_R1 + n_periods, column=5, value=n_periods)
    ws.cell(row=LGD_R1 + n_periods, column=6, value=1.0)
    _style_data(ws.cell(row=LGD_R1 + n_periods, column=6), _PCT_FMT)

    # Check row: VLOOKUP into the E/F LGD lookup table
    lgd_lookup = f'$E${LGD_R1}:$F${LGD_R1 + n_periods}'
    for t in range(n_periods + 1):
        tid_cell = _ref(1, BC + t)  # TID number from header row
        formula = f'=IFERROR(VLOOKUP({tid_cell},{lgd_lookup},2,FALSE),1)'
        cell = ws.cell(row=CHECK_ROW, column=BC + t, value=formula)
        _style_data(cell, _PCT_FMT)

    # ── Per-cohort LGD (column C) — IFERROR HLOOKUP into Check row ──
    last_tid_col = get_column_letter(BC + n_periods)
    for i in range(n_cohorts):
        r = BR1 + i
        a_ref = f'$A{r}'
        lookup_range = f'${get_column_letter(BC)}$1:{last_tid_col}${CHECK_ROW}'
        cell = ws.cell(
            row=r, column=3,
            value=f'=IFERROR(HLOOKUP({a_ref},{lookup_range},{CHECK_ROW},FALSE),1)',
        )
        _style_data(cell, _PCT_FMT)

    # ── EAD in column D (needed for weighted LGD) ──
    for i in range(n_cohorts):
        r = BR1 + i
        _write_val(ws, r, 4, float(eads_window[i]), fmt=_NUM_FMT)

    # ── Weighted LGD — FORMULA ──
    lgd_col = _col_rng(3, BR1, BR2)
    ead_col = _col_rng(4, BR1, BR2)
    cell = ws.cell(
        row=CHECK_ROW, column=3,
        value=f'=IFERROR(SUMPRODUCT({lgd_col},{ead_col})/SUM({ead_col}),"")',
    )
    _style_data(cell, _PCT_FMT)
    # Python cross-check value
    _write_val(ws, CHECK_ROW, 5, f'Python: {weighted_lgd_value:.6f}',
               font=_PARAM_FONT)

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    return ws


def _write_section_header(ws, start_row, title, n_tid, tid_col_start=4):
    """Write a section title + header row (Vintage, Period, Hindsight, TID 0..N).

    Returns the first data row number.
    """
    ws.cell(row=start_row, column=1, value=title).font = _SECTION_FONT
    hdr_row = start_row + 1
    for j, lbl in enumerate(['Vintage', 'Period', 'Hindsight']):
        c = ws.cell(row=hdr_row, column=j + 1, value=lbl)
        _style_hdr(c)
    for t in range(n_tid):
        c = ws.cell(row=hdr_row, column=tid_col_start + t, value=t)
        _style_hdr(c)
    return hdr_row + 1


def _write_section_labels(ws, data_r1, vintage_labels, periods, hindsights):
    """Write vintage/period/hindsight labels for a section's data rows."""
    for i in range(len(vintage_labels)):
        r = data_r1 + i
        _write_val(ws, r, 1, vintage_labels[i])
        p = periods[i]
        if hasattr(p, 'to_pydatetime'):
            p = p.to_pydatetime()
        _write_val(ws, r, 2, p, fmt=_DATE_FMT)
        _write_val(ws, r, 3, hindsights[i])


def _write_section_values(ws, data_r1, matrix, n_tid, tid_col_start=4):
    """Write a matrix of values into a section's data area."""
    n_v = matrix.shape[0]
    for i in range(n_v):
        for t in range(min(n_tid, matrix.shape[1])):
            val = matrix[i, t]
            cell = ws.cell(row=data_r1 + i, column=tid_col_start + t)
            cell.value = None if np.isnan(val) else float(val)
            _style_data(cell, _PCT_FMT)


def _write_master_data_sheet(wb, master_df: pd.DataFrame):
    """Write the RR LGD TERM STRUCTURE ALL sheet from the loaded DataFrame."""
    ws = wb.create_sheet('RR LGD TERM STRUCTURE ALL')

    for col_idx, col_name in enumerate(master_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _style_hdr(cell)

    for row_idx, (_, row) in enumerate(master_df.iterrows(), start=2):
        for col_idx, col_name in enumerate(master_df.columns, start=1):
            val = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            if pd.isna(val):
                cell.value = None
            elif isinstance(val, (np.floating, float)):
                cell.value = float(val)
            elif isinstance(val, (np.integer, int)):
                cell.value = int(val)
            else:
                cell.value = val
            _style_data(cell)

    ws.column_dimensions['A'].width = 14
    return ws


def export_full_audit_workbook(
    master_df: pd.DataFrame,
    scenario: ScenarioResult,
    config: ModelConfig,
    output_path: str | None = None,
) -> bytes:
    """Generate a full audit trail workbook with LIVE Excel formulas.

    Every intermediate calculation (recovery, cumulative balance, discount,
    LGD components, LGD term structure) uses a live formula so auditors can
    click any cell and trace the calculation back to the balance triangle.

    The backtest sheet uses formulas for residuals (=Actual-Forecast),
    mean/std (=AVERAGE/STDEV), and CI bounds (=MIN/MAX with scaling).
    """
    balance_matrix = extract_balance_matrix(master_df)
    eads = master_df['EAD'].values.astype(float)
    master_tids = master_df['TID'].values.astype(float)
    n_total = len(master_df)

    vr_list = scenario.vintage_results
    bt = scenario.backtest
    n_v = len(vr_list)
    # Always use the full max_tid — the window_size controls min_obs_window,
    # NOT the number of TID columns (reference workbooks all have 61 columns)
    n_tid = config.max_tid + 1

    wb = Workbook()
    wb.remove(wb.active)

    # 1. Master data sheet
    _write_master_data_sheet(wb, master_df)

    # 2. LGD Term Structure Summary (values — overview)
    ws_sum = wb.create_sheet('LGD Term Structure Summary')
    _write_val(ws_sum, 1, 1, 'Vintage', font=_HDR_FONT)
    ws_sum.cell(1, 1).fill = _HDR_FILL
    _write_val(ws_sum, 1, 2, 'Period', font=_HDR_FONT)
    ws_sum.cell(1, 2).fill = _HDR_FILL
    for t in range(n_tid):
        cell = ws_sum.cell(row=1, column=3 + t, value=t)
        _style_hdr(cell)
    for i, v in enumerate(vr_list):
        r = 2 + i
        _write_val(ws_sum, r, 1, v.vintage_label)
        period_val = v.period
        if hasattr(period_val, 'to_pydatetime'):
            period_val = period_val.to_pydatetime()
        _write_val(ws_sum, r, 2, period_val, fmt=_DATE_FMT)
        for t in range(min(n_tid, len(v.lgd_term_structure))):
            _write_val(ws_sum, r, 3 + t, float(v.lgd_term_structure[t]), fmt=_PCT_FMT)
    ws_sum.column_dimensions['A'].width = 18
    ws_sum.column_dimensions['B'].width = 12

    # 3. Individual vintage calculation sheets — ALL FORMULAS
    for vi, v in enumerate(vr_list):
        start_idx = v.start_idx
        end_idx = v.end_idx
        n_periods = min(config.max_tid, balance_matrix.shape[1])

        window = balance_matrix[start_idx:end_idx, :n_periods].copy()
        eads_window = eads[start_idx:end_idx].copy()
        tids_window = master_tids[start_idx:end_idx].copy()

        # Apply the observation mask (data preprocessing — stays as values)
        offset = n_total - end_idx
        for i in range(window.shape[0]):
            cohort_master_idx = start_idx + i
            adjusted_max_tid = int(master_tids[cohort_master_idx]) - offset
            if adjusted_max_tid < n_periods:
                window[i, max(0, adjusted_max_tid + 1):] = np.nan

        cohort_indices = np.arange(start_idx, end_idx)

        _write_vintage_calc_sheet(
            wb, v.vintage_label, window, eads_window, tids_window,
            n_periods, config, cohort_indices, v.weighted_lgd,
        )

    # 4. LGD Backtest Summary — FORMULAS for residuals, mean/std, CI
    ws_bt = wb.create_sheet('LGD Backtest Summary')
    vintage_labels = bt.vintage_labels
    periods = bt.periods
    hindsights = list(range(n_v, 0, -1))
    TC = 4  # TID columns start at column D

    # ── FORECAST (values) ──
    fc_r1 = _write_section_header(ws_bt, 1, 'FORECAST LGD BY VINTAGE', n_tid)
    _write_section_labels(ws_bt, fc_r1, vintage_labels, periods, hindsights)
    _write_section_values(ws_bt, fc_r1, bt.forecast_matrix, n_tid)
    fc_r2 = fc_r1 + n_v - 1

    # ── ACTUAL (values) ──
    ac_start = fc_r2 + 3
    ac_r1 = _write_section_header(ws_bt, ac_start, 'ACTUAL LGD BY VINTAGE (DIAGONAL)', n_tid)
    _write_section_labels(ws_bt, ac_r1, vintage_labels, periods, hindsights)
    _write_section_values(ws_bt, ac_r1, bt.actual_matrix, n_tid)
    ac_r2 = ac_r1 + n_v - 1

    # ── MEAN / STD — FORMULAS ──
    ms_start = ac_r2 + 3
    ws_bt.cell(row=ms_start, column=1, value='MEAN / STD DEVIATION').font = _SECTION_FONT
    MEAN_ROW = ms_start + 1
    STD_ROW = ms_start + 2
    _write_val(ws_bt, MEAN_ROW, 1, 'Mean', font=_LABEL_FONT)
    _write_val(ws_bt, STD_ROW, 1, 'Std Deviation', font=_LABEL_FONT)
    for t in range(n_tid):
        col = TC + t
        fc_col = _col_rng(col, fc_r1, fc_r2)
        cell = ws_bt.cell(row=MEAN_ROW, column=col, value=f'=AVERAGE({fc_col})')
        _style_data(cell, _PCT_FMT)
        cell = ws_bt.cell(row=STD_ROW, column=col, value=f'=STDEV({fc_col})')
        _style_data(cell, _PCT_FMT)

    # ── MEAN MATRIX — FORMULAS (mean where actual exists) ──
    mm_start = STD_ROW + 2
    mm_r1 = _write_section_header(ws_bt, mm_start, 'MEAN LGD BY VINTAGE', n_tid)
    _write_section_labels(ws_bt, mm_r1, vintage_labels, periods, hindsights)
    for i in range(n_v):
        for t in range(n_tid):
            col = TC + t
            ac_cell = _ref(ac_r1 + i, col)
            mean_cell = _ref(MEAN_ROW, col)
            cell = ws_bt.cell(
                row=mm_r1 + i, column=col,
                value=f'=IF(ISBLANK({ac_cell}),"",{mean_cell})',
            )
            _style_data(cell, _PCT_FMT)
    mm_r2 = mm_r1 + n_v - 1

    # ── CI PARAMETERS ──
    z_score = bt.z_score
    ci_pctl = bt.ci_percentile
    PARAM_ROW_BT = mm_r2 + 2
    _write_val(ws_bt, PARAM_ROW_BT, 1, 'z-score', font=_PARAM_FONT)
    _write_val(ws_bt, PARAM_ROW_BT, 2, z_score)
    _write_val(ws_bt, PARAM_ROW_BT, 3, 'CI percentile', font=_PARAM_FONT)
    _write_val(ws_bt, PARAM_ROW_BT, 4, ci_pctl)
    _write_val(ws_bt, PARAM_ROW_BT, 5, 'TID cap (n_v-1)', font=_PARAM_FONT)
    _write_val(ws_bt, PARAM_ROW_BT, 6, n_v - 1)

    z_ref = _abs_ref(PARAM_ROW_BT, 2)
    nv_cap_ref = _abs_ref(PARAM_ROW_BT, 6)

    # ── UPPER CI — FORMULAS ──
    # Upper(i,t) = MIN(1, Forecast_oldest[t] + z * Std[t] * SQRT(h / MIN(t, cap)))
    up_start = PARAM_ROW_BT + 2
    up_r1 = _write_section_header(
        ws_bt, up_start,
        f'UPPER BOUND ({ci_pctl*100:.0f}th pctl) \u2014 Binomial CI',
        n_tid,
    )
    _write_section_labels(ws_bt, up_r1, vintage_labels, periods, hindsights)
    for i in range(n_v):
        h = i + 1
        start_t = h  # staircase: CI starts at TID = h
        for t in range(n_tid):
            col = TC + t
            cell = ws_bt.cell(row=up_r1 + i, column=col)
            if i >= n_v - 1 or t < start_t or t == 0:
                cell.value = None
            else:
                fc_oldest = _ref(fc_r1, col)  # Oldest vintage forecast
                std_cell = _ref(STD_ROW, col)
                cell.value = (
                    f'=MIN(1,{fc_oldest}+{z_ref}*{std_cell}'
                    f'*SQRT({h}/MIN({t},{nv_cap_ref})))'
                )
            _style_data(cell, _PCT_FMT)
    up_r2 = up_r1 + n_v - 1

    # ── LOWER CI — FORMULAS ──
    lo_start = up_r2 + 3
    lo_r1 = _write_section_header(
        ws_bt, lo_start,
        f'LOWER BOUND ({(1-ci_pctl)*100:.0f}th pctl) \u2014 Binomial CI',
        n_tid,
    )
    _write_section_labels(ws_bt, lo_r1, vintage_labels, periods, hindsights)
    for i in range(n_v):
        h = i + 1
        start_t = h
        for t in range(n_tid):
            col = TC + t
            cell = ws_bt.cell(row=lo_r1 + i, column=col)
            if i >= n_v - 1 or t < start_t or t == 0:
                cell.value = None
            else:
                fc_oldest = _ref(fc_r1, col)
                std_cell = _ref(STD_ROW, col)
                cell.value = (
                    f'=MAX(0,{fc_oldest}-{z_ref}*{std_cell}'
                    f'*SQRT({h}/MIN({t},{nv_cap_ref})))'
                )
            _style_data(cell, _PCT_FMT)
    lo_r2 = lo_r1 + n_v - 1

    # ── RESIDUALS — FORMULAS (= Actual - Forecast) ──
    res_start = lo_r2 + 3
    res_r1 = _write_section_header(
        ws_bt, res_start, 'DIFFERENCE (ACTUAL - FORECAST)', n_tid,
    )
    _write_section_labels(ws_bt, res_r1, vintage_labels, periods, hindsights)
    for i in range(n_v):
        for t in range(n_tid):
            col = TC + t
            ac = _ref(ac_r1 + i, col)
            fc = _ref(fc_r1 + i, col)
            cell = ws_bt.cell(
                row=res_r1 + i, column=col,
                value=f'=IF(ISBLANK({ac}),"",{ac}-{fc})',
            )
            _style_data(cell, _PCT_FMT)
    res_r2 = res_r1 + n_v - 1

    # ── Summary statistics ──
    row = res_r2 + 3
    ws_bt.cell(row=row, column=1, value='BACKTEST SUMMARY STATISTICS').font = _SECTION_FONT
    row += 1
    stats = [
        ('Window Size (months)', scenario.window_size),
        ('Number of Vintages', n_v),
        ('Number of Residuals', scenario.n_residuals),
        ('Discount Rate', config.discount_rate),
        ('CI Percentile', ci_pctl),
        ('z-score', z_score),
        ('', ''),
        ('Mean Error (Bias)', scenario.mean_error),
        ('Std Dev of Errors', scenario.std_error),
        ('RMSE', scenario.rmse),
        ('MAE', scenario.mae),
        ('Median AE', scenario.median_ae),
        ('Max |Error|', scenario.max_abs_error),
        ('IQR', scenario.iqr_error),
        ('Composite Score', scenario.composite_score),
        ('', ''),
        ('JB Rejects Normality', scenario.jb_reject),
        ('Overall CI Coverage', bt.overall_coverage),
        ('Latest LGD at TID=0', scenario.latest_lgd_tid0),
        ('Latest Weighted LGD', scenario.latest_weighted_lgd),
    ]
    for label, val in stats:
        _write_val(ws_bt, row, 1, label, font=_LABEL_FONT)
        cell = _write_val(ws_bt, row, 2, val)
        if isinstance(val, float) and abs(val) < 2:
            cell.number_format = '0.000000'
        row += 1

    ws_bt.column_dimensions['A'].width = 22
    ws_bt.column_dimensions['B'].width = 14
    ws_bt.column_dimensions['C'].width = 14

    # Hide gridlines on all sheets
    for ws in wb.worksheets:
        _hide_gridlines(ws)

    # Serialize to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    workbook_bytes = buf.getvalue()

    if output_path is not None:
        with open(output_path, 'wb') as f:
            f.write(workbook_bytes)

    return workbook_bytes


def export_all_audit_workbooks_zip(
    master_df: pd.DataFrame,
    scenarios: list[ScenarioResult],
    config: ModelConfig,
) -> bytes:
    """Generate a ZIP containing full audit trail workbooks for all scenarios."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for s in scenarios:
            w = s.window_size
            cfg = ModelConfig(
                discount_rate=config.discount_rate,
                window_size=config.window_size,
                max_tid=config.max_tid,
                lgd_cap=config.lgd_cap,
                ci_percentile=config.ci_percentile,
                min_obs_window=w if w < config.window_size else None,
            )
            wb_bytes = export_full_audit_workbook(master_df, s, cfg)
            zf.writestr(f'LGD_Audit_{w}m.xlsx', wb_bytes)
    buf.seek(0)
    return buf.getvalue()
